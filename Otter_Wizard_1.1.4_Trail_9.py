import tkinter as tk
from tkinter import filedialog, ttk, simpledialog, messagebox
import pandas as pd
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import logging
import json
import pygame

class OtterWizard:
    def __init__(self, root):
        self.root = root
        self.root.title("Otter Wizard")

        # Set up logging
        logging.basicConfig(filename='Otter_wizard_log.log', level=logging.INFO, format='%(asctime)s:%(levelname)s:%(message)s')

        # Default start row and column
        self.start_row = 1
        self.start_col = 1

        # Determine the path to store the config file
        self.config_file_path = os.path.join(os.path.dirname(__file__), 'config.json')

        # Load user preferences
        self.load_user_preferences()

        # Create UI elements
        self.create_widgets()

        # Display version info in the corner
        self.version_label = tk.Label(self.root, text="Version 1.1.3", anchor="se")
        self.version_label.grid(row=7, column=2, sticky="se", padx=10, pady=10)  # Moved to row 7

    def load_user_preferences(self):
        try:
            with open(self.config_file_path, 'r') as f:
                config = json.load(f)
                self.users = config.get('users', ['Casey', 'Darvis', 'Otter'])  # Default users if not found
                self.default_user = config.get('default_user', self.users[0])  # Default user if not found
                self.complete_sound_path = config.get('complete_sound_path', '')  # Complete sound file path
                self.play_complete_sound_on_success = config.get('play_complete_sound_on_success', True)  # Play complete sound on success
        except FileNotFoundError:
            self.users = ['Casey', 'Darvis', 'Otter']  # Default users if config file doesn't exist
            self.default_user = self.users[0]  # Default user if config file doesn't exist
            self.complete_sound_path = ''  # Default complete sound file path
            self.play_complete_sound_on_success = True  # Default to play complete sound on success

    def save_user_preferences(self):
        config = {
            'users': self.users,
            'default_user': self.default_user,
            'complete_sound_path': self.complete_sound_path,
            'play_complete_sound_on_success': self.play_complete_sound_on_success
        }
        with open(self.config_file_path, 'w') as f:
            json.dump(config, f)

    def create_widgets(self):
        # Station Import File selection
        self.station_label = tk.Label(self.root, text="Station Import File:")
        self.station_label.grid(row=0, column=0, padx=10, pady=10)
        self.station_entry = tk.Entry(self.root, width=50)
        self.station_entry.grid(row=0, column=1, padx=10, pady=10)
        self.station_button = tk.Button(self.root, text="Browse", command=self.browse_station, width=32)
        self.station_button.grid(row=0, column=2, columnspan=3, padx=10, pady=10)

        # Glossary Template selection
        self.glossary_label = tk.Label(self.root, text="Glossary Template:")
        self.glossary_label.grid(row=1, column=0, padx=10, pady=10)
        self.glossary_entry = tk.Entry(self.root, width=50)
        self.glossary_entry.grid(row=1, column=1, padx=10, pady=10)
        self.glossary_button = tk.Button(self.root, text="Browse", command=self.browse_glossary, width=32)
        self.glossary_button.grid(row=1, column=2, columnspan=3, padx=10, pady=10)

        # Output Location and Name
        self.output_label = tk.Label(self.root, text="Output Location and Name:")
        self.output_label.grid(row=2, column=0, padx=10, pady=10)
        self.output_entry = tk.Entry(self.root, width=50)
        self.output_entry.grid(row=2, column=1, padx=10, pady=10)
        self.output_button = tk.Button(self.root, text="Browse", command=self.browse_output, width=32)
        self.output_button.grid(row=2, column=2, columnspan=3, padx=10, pady=10)

        # User selection
        self.user_label = tk.Label(self.root, text="User:")
        self.user_label.grid(row=3, column=0, padx=10, pady=10)
        self.user_var = tk.StringVar(value=self.default_user)
        self.user_entry = ttk.Combobox(self.root, textvariable=self.user_var, values=self.users, width=48)  # Adjusted width
        self.user_entry.grid(row=3, column=1, padx=10, pady=10)
        self.user_button_frame = tk.Frame(self.root)
        self.user_button_frame.grid(row=3, column=2, padx=10, pady=10)

        self.user_add_button = tk.Button(self.user_button_frame, text="Add", command=self.add_user, width=8)
        self.user_add_button.pack(side=tk.LEFT, padx=5)

        self.user_remove_button = tk.Button(self.user_button_frame, text="Remove", command=self.remove_user, width=8)
        self.user_remove_button.pack(side=tk.LEFT, padx=5)

        self.user_default_button = tk.Button(self.user_button_frame, text="Set Default", command=self.set_default_user, width=10)
        self.user_default_button.pack(side=tk.LEFT, padx=5)

        # Run button
        self.run_button = tk.Button(self.root, text="Run", command=self.run, width=44)
        self.run_button.grid(row=4, column=1, padx=10, pady=10)

        # Progress Bar
        self.progress_label = tk.Label(self.root, text="Progress:")
        self.progress_label.grid(row=5, column=0, padx=10, pady=10)
        self.progress = ttk.Progressbar(self.root, orient='horizontal', mode='determinate', length=600)
        self.progress.grid(row=5, column=1, columnspan=4, padx=10, pady=10)

        # Help Button
        self.help_button = tk.Button(self.root, text="Help", command=self.show_help, width=32)
        self.help_button.grid(row=4, column=2, columnspan=3, padx=10, pady=10)

        # Complete Sound Settings
        self.complete_sound_label = tk.Label(self.root, text="Complete Sound File Name:")
        self.complete_sound_label.grid(row=6, column=0, padx=10, pady=10)
        self.complete_sound_entry = tk.Entry(self.root, width=50)
        self.complete_sound_entry.grid(row=6, column=1, padx=10, pady=10)
        self.complete_sound_button = tk.Button(self.root, text="Select MP3 File", command=self.browse_complete_sound, width=32)
        self.complete_sound_button.grid(row=6, column=2, columnspan=3, padx=10, pady=10)

        self.play_complete_sound_var = tk.BooleanVar(value=self.play_complete_sound_on_success)
        self.play_complete_sound_checkbox = tk.Checkbutton(self.root, text="Play Complete Sound on Success", variable=self.play_complete_sound_var)
        self.play_complete_sound_checkbox.grid(row=7, column=1, padx=10, pady=10)

        # Display default complete sound path if it exists
        if self.complete_sound_path:
            self.complete_sound_entry.delete(0, tk.END)
            self.complete_sound_entry.insert(0, os.path.basename(self.complete_sound_path))

    def browse_station(self):
        station_file = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if station_file:
            self.station_entry.delete(0, tk.END)
            self.station_entry.insert(0, station_file)

    def browse_glossary(self):
        glossary_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if glossary_file:
            self.glossary_entry.delete(0, tk.END)
            self.glossary_entry.insert(0, glossary_file)

    def browse_output(self):
        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_file:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, output_file)

    def browse_complete_sound(self):
        complete_sound_file = filedialog.askopenfilename(filetypes=[("MP3 files", "*.mp3")])
        if complete_sound_file:
            self.complete_sound_entry.delete(0, tk.END)
            self.complete_sound_entry.insert(0, os.path.basename(complete_sound_file))  # Display only the filename in the entry
            self.complete_sound_path = complete_sound_file  # Store the full path
            self.save_user_preferences()

    def add_user(self):
        new_user = simpledialog.askstring("Add User", "Enter new user name:")
        if new_user and new_user not in self.users:
            self.users.append(new_user)
            self.user_entry['values'] = self.users
            self.user_entry.set(new_user)
            logging.info(f"User '{new_user}' added by {self.user_var.get()}.")
            self.save_user_preferences()

    def remove_user(self):
        user_to_remove = simpledialog.askstring("Remove User", "Enter user name to remove:")
        if user_to_remove in self.users and user_to_remove != self.default_user:
            self.users.remove(user_to_remove)
            self.user_entry['values'] = self.users
            self.user_entry.set(self.users[0] if self.users else "")
            logging.info(f"User '{user_to_remove}' removed by {self.user_var.get()}.")
            self.save_user_preferences()
        elif user_to_remove == self.default_user:
            messagebox.showinfo("Cannot Remove", "Cannot remove default user.")
        else:
            messagebox.showinfo("User Not Found", f"User '{user_to_remove}' not found.")

    def set_default_user(self):
        self.default_user = self.user_var.get()
        self.user_var.set(self.default_user)
        logging.info(f"Default user set to: {self.default_user} by {self.user_var.get()}.")
        self.save_user_preferences()

    def run(self):
        station_file = self.station_entry.get()
        glossary_template = self.glossary_entry.get()
        output_file = self.output_entry.get()

        if not station_file or not glossary_template or not output_file:
            messagebox.showerror("Error", "Please fill in all fields")
            logging.error("Missing input fields")
            return

        try:
            self.process_files(station_file, glossary_template, output_file)
            self.show_success_message(station_file, glossary_template, output_file)

            # Play complete sound if enabled
            if self.play_complete_sound_on_success and self.complete_sound_path:
                self.play_complete_sound()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
            logging.error(f"An error occurred: {e} - Process run by {self.user_var.get()}")

    def show_success_message(self, station_file, glossary_template, output_file):
        success_message = f"Process completed successfully! Files: {os.path.basename(station_file)} - {os.path.basename(glossary_template)} - {os.path.basename(output_file)}"
        success_window = tk.Toplevel(self.root)
        success_window.title("Success")
        success_label = tk.Label(success_window, text=success_message)
        success_label.pack(padx=20, pady=20)

        # Close the success window automatically after 3 seconds
        success_window.after(3000, success_window.destroy)

        logging.info(f"Process completed by {self.user_var.get()}: {success_message}")

    def play_complete_sound(self):
        try:
            pygame.mixer.init()
            pygame.mixer.music.load(self.complete_sound_path)
            pygame.mixer.music.play()
            while pygame.mixer.music.get_busy():
                pygame.time.Clock().tick(10)
        except pygame.error as e:
            messagebox.showerror("Complete Sound Error", f"Error playing complete sound: {e}")
            logging.error(f"Error playing complete sound: {e}")

    def process_files(self, station_file, glossary_template, output_file):
        # Update progress bar
        self.progress['value'] = 10
        self.root.update_idletasks()

        # Copy the glossary template
        shutil.copy(glossary_template, output_file)
        self.progress['value'] = 30
        self.root.update_idletasks()

        # Load the new Excel file
        wb = openpyxl.load_workbook(output_file)
        self.progress['value'] = 50
        self.root.update_idletasks()

        # Create a default sheet named "Input" if it doesn't exist
        if "Input" not in wb.sheetnames:
            ws = wb.create_sheet("Input")
        else:
            ws = wb["Input"]

        # Read the CSV data
        station_df = pd.read_csv(station_file)
        self.progress['value'] = 70
        self.root.update_idletasks()

        # Write CSV data to the Excel sheet starting at cell A1
        for r_idx, row in enumerate(dataframe_to_rows(station_df, index=False, header=True), start=self.start_row):
            for c_idx, value in enumerate(row, start=self.start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Save the modified Excel file
        wb.save(output_file)
        self.progress['value'] = 90
        self.root.update_idletasks()

        # Open the saved Excel file
        os.startfile(output_file)
        self.progress['value'] = 100
        self.root.update_idletasks()

    def show_help(self):
        help_text = ("1. Select the Station Import File (CSV format).\n"
                     "2. Select the Glossary Template (Excel format).\n"
                     "3. Choose the output location and filename for the processed Excel file (A named folder is recommended).\n"
                     "4. Click 'Run' to start the process. Progress bar indicates run stage.\n\n"
                     "Ensure all fields are filled before running the process. \n\n"
                     "O.T.T.E.R - Optimized Template-based Transformation for Excel Reports.")
        messagebox.showinfo("Help", help_text)

if __name__ == "__main__":
    root = tk.Tk()
    app = OtterWizard(root)
    root.mainloop()
